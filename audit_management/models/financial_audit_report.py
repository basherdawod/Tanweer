import openpyxl
import io
import base64
from odoo import api, fields, models, _, tools, Command
from odoo.exceptions import AccessError, ValidationError, UserError
from datetime import datetime, date

class FinancialAuditReporting(models.Model):
    _name = "financial.audit.customer"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = "Customer Registration"

    lable1 = fields.Char(
        string="Text", readonly=True,
        default="MODULAR CONCEPTS L.L.C.\n DUBAI - UNITED ARAB EMIRATES \n FINANCIAL STATEMENTS & REPORTS")
    comprehensive_income_ids = fields.One2many(
        'comprehensive.income',
        'financial_id',
    )
    comprehensive_income_line_ids = fields.One2many(
        'comprehensive.income.line',
        'financial_line_id',
    )
    name = fields.Char(
        string="Registration No", readonly=True,
        default=lambda self: _('New'), copy=False)
    partner_id = fields.Many2one('res.partner', string="Customer Name")
    integration_type = fields.Selection(
        string='Integration Type',
        selection=[('current_system', 'Current Systems'),
                   ('customers_system', 'Customers Systems')],
        required=False, default='current_system')
    data_last_years_end = fields.Date(
        string='Registration Date',
        required=True,
    )


    assets_category_ids = fields.One2many(
        comodel_name='account.assets.audit',
        inverse_name='financial_audit_customer_id',
        string='Assets_category_ids',
        required=False)


    upload_xlsx = fields.Binary(string="Upload XLSX File")
    upload_xlsx_filename = fields.Char(string="Filename")



    data_fis_years_end = fields.Date(
        string='Fiscal Year End',
        required=False,
        default=lambda self: datetime(date.today().year, 12, 31).strftime("%Y-%m-%d")
    )
    data_last_years = fields.Date(
        string='Last Fiscal Year End',
        required=False,
        default=lambda self: datetime(date.today().year - 1, 12, 31).strftime("%Y-%m-%d")
    )
    active_audit = fields.Boolean(
        string='Active Account',
        required=False)

    # audit_financial_program_ids = fields.One2many(
    #     comodel_name='audit.financial.program',
    #     inverse_name='partner_id',
    #     string='Customers Audit Report',
    #     required=False)

    audit_char_account_id = fields.Many2one(
        comodel_name='audit.account.account',
        string='Chart Of Account',
        required=False)
    account_lines_ss = fields.One2many(
        comodel_name='audit.account.account.line',
        inverse_name='account_ids_audit1',
        string='Account lines',
        required=False)
    account_type_level = fields.Many2one('account.type.level', string="Account Type")






    #
    # def action_import_account_lines(self):
    #     if not self.upload_xlsx:
    #         raise ValidationError(_("Please upload an XLSX file first."))
    #
    #     try:
    #         file_content = base64.b64decode(self.upload_xlsx)
    #         workbook = xlrd.open_workbook(file_contents=file_content)
    #         sheet = workbook.sheet_by_index(0)
    #
    #         lines = []
    #         for row_idx in range(1, sheet.nrows):
    #             code = sheet.cell_value(row_idx, 0)
    #             account_type = sheet.cell_value(row_idx, 1)
    #             opening_balance = sheet.cell_value(row_idx, 2)
    #
    #             lines.append((0, 0, {
    #                 'code': code,
    #                 'account_name': account_name,
    #                 'account_balance': opening_balance,
    #
    #             }))
    #
    #         self.account_lines_ss = lines
    #
    #     except Exception as e:
    #         raise ValidationError(_("Error processing the XLSX file: %s") % str(e))
    def action_import_account_lines(self):
        if not self.upload_xlsx:
            raise ValidationError(_("Please upload an XLSX file first."))

        try:
            # Decode the uploaded XLSX file content
            file_content = base64.b64decode(self.upload_xlsx)
            file = io.BytesIO(file_content)

            # Load the workbook and get the active sheet
            workbook = openpyxl.load_workbook(filename=file)
            sheet = workbook.active

            # List of valid account types
            valid_account_types = [
                "asset_receivable", "asset_cash", "asset_current", "asset_non_current",
                "asset_prepayments", "asset_fixed", "liability_payable",
                "liability_credit_card", "liability_current", "liability_non_current",
                "equity", "equity_unaffected", "income", "income_other", "expense",
                "expense_depreciation", "expense_direct_cost", "off_balance"
            ]

            # Initialize an empty list for lines
            lines = []

            # Iterate through the rows, starting from row 2 (assuming row 1 is headers)
            for row_idx in range(2, sheet.max_row + 1):
                # Read data from each cell
                code = sheet.cell(row=row_idx, column=1).value
                account_name = sheet.cell(row=row_idx, column=2).value
                account_type = sheet.cell(row=row_idx, column=3).value
                opening = sheet.cell(row=row_idx, column=4).value or 0
                opening_debit = sheet.cell(row=row_idx, column=5).value or 0
                opening_credit = sheet.cell(row=row_idx, column=6).value or 0
                opening_balance = sheet.cell(row=row_idx, column=7).value or 0

                # Validate required fields
                if not code or not account_name or not account_type:
                    raise ValidationError(
                        _("Row %d: 'Code', 'Account Name', and 'Account Type' are required.") % row_idx)

                # Validate account_type
                if account_type not in valid_account_types:
                    raise ValidationError(_("Row %d: Invalid 'Account Type' value '%s'. Must be one of: %s") %
                                          (row_idx, account_type, ", ".join(valid_account_types)))

                # Append the line to the list
                lines.append((0, 0, {
                    'code': code,
                    'name': account_name,
                    'account_type': account_type,
                    'opening_balance': opening,
                    'opening_debit': opening_debit,
                    'opening_credit': opening_credit,
                    'current_balance': opening_balance,
                }))

            # Assign the processed lines to the One2many field
            self.account_lines_ss = lines

        except Exception as e:
            raise ValidationError(_("Error processing the XLSX file: %s") % str(e))

    

    # def action_import_account_lines(self):
    #     if not self.upload_xlsx:
    #         raise ValidationError(_("Please upload an XLSX file first."))

    #     try:
    #         file_content = base64.b64decode(self.upload_xlsx)

    #         file = io.BytesIO(file_content)
    #         workbook = openpyxl.load_workbook(filename=file) 
    #         sheet = workbook.active
            
    #         lines = []
    #         for row_idx in range(1, sheet.max_row): 
    #             code = sheet.cell(row=row_idx + 1, column=1).value
    #             account_name = sheet.cell(row=row_idx + 1, column=2).value or 0
    #             account_type = sheet.cell(row=row_idx + 1, column=3).value or 0 
    #             opening = sheet.cell(row=row_idx + 1, column=4).value or 0 
    #             opening_debit = sheet.cell(row=row_idx + 1, column=5).value or 0 
    #             opening_credit = sheet.cell(row=row_idx + 1, column=6).value or 0  
    #             opening_balance = sheet.cell(row=row_idx + 1, column=7).value or 0  
                
    #             lines.append((0, 0, {
    #                 'code': code,
    #                 'account_name': account_name,
    #                 'account_type': account_type,
    #                 'opening': opening,
    #                 'opening_debit': opening_debit,
    #                 'opening_credit': opening_credit,
    #                 'account_balance': opening_balance,
    #             }))
            
    #         self.account_lines_ss = lines

    #     except Exception as e:
    #         raise ValidationError(_("Error processing the XLSX file: %s") % str(e))

    def create_account_lines_customers(self):
        if self.integration_type == 'current_system':
            self.account_lines_ss.unlink()
            self.assets_category_ids.unlink()
            list_account = []
            list_category = []
            for record in self :
                years1 = record.data_fis_years_end.year
                start_this_years = datetime(years1, 1, 31).strftime("%Y-%m-%d")
                years2 = record.data_last_years.year
                start_last_years =datetime(years2, 1, 31).date()
                years_2last = datetime(years2-1, 12, 31).date()
                yy = years_2last.year
                start_2_last = datetime(yy, 1, 31).date()
                account_audit_customers = self.env['audit.account.account.line'].search([('account_ids_audit1', '=', record.id)])
                lines = self.env['account.account'].search([])
                account_asset = self.env['account.asset.asset'].search([])
                for rec in account_asset :
                    for cat in rec.category_id :
                        for account in cat.account_asset_id:
                            list_category.append({
                            'code': account.code,
                            'account_name': account.name,
                            'name': cat.name,
                            'asset_name': rec.name,
                            'date': rec.date,
                            'cross_value': rec.value,
                            'residual': rec.value_residual,
                            'cumulate':rec.value - rec.value_residual ,
                        })
                if record.id not in account_audit_customers.ids:
                    for account in lines :
                        total_credit = 0.0
                        total_debit = 0.0
                        total_balance = 0.0
                        open_balance = 0.0
                        balance_2years = 0.0

                        move_lines = self.env['account.move.line'].search([
                            ('account_id', '=', account.id)
                            ])

                        for mov in move_lines:
                            if  mov.date <= record.data_fis_years_end :
                                total_credit += mov.credit
                                total_debit += mov.debit
                                total_balance += mov.debit - mov.credit
                            if start_last_years <= mov.date <= record.data_last_years :
                                open_balance += mov.debit - mov.credit
                            if start_2_last  <= mov.date <= years_2last :
                                balance_2years += mov.debit - mov.credit
                                # balance_2years += mov.debit - mov.credit
                        if record.active_audit :
                            if total_credit or total_balance or open_balance or balance_2years or total_debit != 0.0:
                                list_account.append({
                                'code': account.code,
                                'name': account.name,
                                'account_type': account.account_type,
                                'current_balance': total_balance,
                                'opening_balance': open_balance,
                                'balance_2years': balance_2years,
                                'opening_credit': total_credit,
                                'opening_debit': total_debit,
                            })
                        else:
                            list_account.append({
                                'code': account.code,
                                'name': account.name,
                                'account_type': account.account_type,
                                'current_balance': total_balance,
                                'opening_balance': open_balance,
                                'balance_2years': balance_2years,
                                'opening_credit': total_credit,
                                'opening_debit': total_debit,
                            })

                record.write({
                    'account_lines_ss': [Command.create(vals) for vals in list_account] ,
                    'assets_category_ids': [Command.create(vals) for vals in list_category]
                })


    def write(self, vals):
        # Check if the field to be updated is already in the vals dictionary
        if 'account_lines_ss' not in vals:  # Only modify if not already in vals
            for record in self:
                if record.integration_type != 'customers_system':  # Example condition
                    vals['account_lines_ss'] = None  # Setting the field to None
        # Call the parent method to actually update the record
        return super(FinancialAuditReporting, self).write(vals)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('financial.audit.customer') or _('New')
        records = super(FinancialAuditReporting, self).create(vals_list)
        return records


    def create_audit_report(self):
        for record in self:
            existing_audit_reports = record.audit_financial_program_ids
            if not existing_audit_reports:
                audit_report_name = record.name
            else:
                next_letter = chr(65 + len(existing_audit_reports))  # ASCII 'A' is 65
                audit_report_name = f"{record.name}/{next_letter}"

            # Create the new audit_report record
            self.env['audit.financial.program'].create({
                'partner_id': record.id,
                'name': audit_report_name,
            })

    def create_comprehensive_report(self):
        for record in self:
            existing_comprehensive_report = record.comprehensive_income_ids
            if not existing_comprehensive_report:
                audit_report_name = record.name
            else:
                next_letter = chr(65 + len(existing_comprehensive_report))  # ASCII 'A' is 65
                audit_report_name = f"{record.name}/{next_letter}"

            # Create the new audit_report record
            self.env['comprehensive.income'].create({
                'partner_id':record.id,
                'name': audit_report_name,
            })

    # def create_comprehensive_report(self):
    #     for record in self:
    #         existing_comprehensive_report = record.comprehensive_income_ids
    #         if not existing_comprehensive_report:
    #             audit_report_name = record.name
    #         else:
    #             next_letter = chr(65 + len(existing_comprehensive_report))  # ASCII 'A' is 65
    #             audit_report_name = f"{record.name}/{next_letter}"
    #
    #         # Create the new audit_report record
    #         new_report = self.env['comprehensive.income'].create({
    #             'financial_id': record.id,
    #             'name': audit_report_name,
    #         })
    #
    #         for line in record.comprehensive_income_line_ids:
    #             self.env['comprehensive.income.line'].create({
    #                 'comprehensive_income_id': new_report.id,
    #                 'code': line.code,
    #                 'account_name': line.name,
    #                 'total_last_year': line.opening_balance,
    #                 'total_this_year': line.current_balance,
    #
    #             })


class AuditAccountChar(models.Model):
    _name = 'audit.account.account' #model_audit_account_account
    _description = 'AuditAccountChar'
    _inherit = ['mail.thread', 'mail.activity.mixin']


    name = fields.Char(string="Account Name", required=True, index='trigram', tracking=True, translate=True)
    customer_account_id = fields.Many2one(
        comodel_name='financial.audit.customer',
        string='Customer Account ',
        required=False)
    account_lines_ids = fields.One2many(
        comodel_name='audit.account.account.line',
        inverse_name='account_ids_audit',
        string='Account_lines_ids',
        required=False)



    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('audit.account.account') or _('New')
        records = super(AuditAccountChar, self).create(vals_list)
        return records


class AuditAccountCharLine(models.Model):
    _name = 'audit.account.account.line' #model_audit_account_account_line
    _description = 'audit.account.account.line'


    name = fields.Char(string="Account Name", required=True, index='trigram', tracking=True, translate=True)
    account_ids_audit = fields.Many2one(
        comodel_name='audit.account.account',
        string='Account_ids_audit',
        required=False)
    account_ids_audit1 = fields.Many2one(
        comodel_name='financial.audit.customer',
        string='Account_ids_audit',
        required=False)
    code = fields.Char(size=64, required=True, tracking=True, index=True, unaccent=False)
    account_type = fields.Selection(
        selection=[
            ("asset_receivable", "Receivable"),
            ("asset_cash", "Bank and Cash"),
            ("asset_current", "Current Assets"),
            ("asset_non_current", "Non-current Assets"),
            ("asset_prepayments", "Prepayments"),
            ("asset_fixed", "Fixed Assets"),
            ("liability_payable", "Payable"),
            ("liability_credit_card", "Credit Card"),
            ("liability_current", "Current Liabilities"),
            ("liability_non_current", "Non-current Liabilities"),
            ("equity", "Equity"),
            ("equity_unaffected", "Current Year Earnings"),
            ("income", "Income"),
            ("income_other", "Other Income"),
            ("expense", "Expenses"),
            ("expense_depreciation", "Depreciation"),
            ("expense_direct_cost", "Cost of Revenue"),
            ("off_balance", "Off-Balance Sheet"),
        ],
        string="Type", tracking=True,
        required=True,
        store=True, readonly=False, precompute=True, index=True,
        help="Account Type is used for information purpose, to generate country-specific legal reports, and set the rules to close a fiscal year and generate opening entries."
    )

    opening_debit = fields.Float(string="Debit" )
    opening_credit = fields.Float(string="Credit"  )

    current_balance = fields.Float(string="Current balance")

    opening_balance = fields.Float(string="Balance before 1 years")
    balance_2years = fields.Float(string="Balance before 2 years")


class AccountAssetsAudit(models.Model):
    _name = 'account.assets.audit'
    _description = 'AccountAssetsAudit'

    name = fields.Char(string="Assets Name")
    asset_name = fields.Char(string="Assets Category")

    financial_audit_customer_id= fields.Many2one(
        comodel_name='financial.audit.customer',
        string='Customer Assets Category',
        required=False)
    # assets_category = fields.Many2one(
    #     comodel_name='account.asset.asset',
    #     string='Assets_category',
    #     required=False)
    cross_value = fields.Float(
        string='Cross Value',
        required=False)
    date = fields.Date(
        string='Date',
        required=False)
    residual = fields.Float(
        string='Residual',
        required=False)
    cumulate = fields.Float(
        string='cumulate',
        required=False)


    code = fields.Char(size=64, required=True, tracking=True, index=True, unaccent=False)
    account_name = fields.Char(string="Account Name", required=True, index='trigram', tracking=True, translate=True)