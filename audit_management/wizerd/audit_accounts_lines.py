# from odoo import models, fields, api
# import base64
# import io
# import pandas as pd

# class AuditAccountImportWizard(models.Model):
#     _name = 'audit.account.import.wizard'
#     _description = 'Import Wizard'

#     name = fields.Char(string='Name')
#     upload_file = fields.Binary(string="Upload Excel File")
#     upload_filename = fields.Char(string="File Name")
#     one2many_line_ids = fields.One2many('your.model.line', 'parent_id', string="Lines")

#     def action_import_lines(self):
#         """Read the xlsx file and populate the One2many lines."""
#         if not self.upload_file:
#             raise ValueError("Please upload an Excel file.")
        
#         # Decode the uploaded file
#         file_content = base64.b64decode(self.upload_file)
#         file_data = io.BytesIO(file_content)
        
#         try:
#             # Read the Excel file using pandas
#             df = pd.read_excel(file_data)
            
#             # Iterate through the rows and create One2many lines
#             lines = []
#             for _, row in df.iterrows():
#                 lines.append((0, 0, {
#                     'field1': row.get('Field1'),
#                     'field2': row.get('Field2'),
#                     'field3': row.get('Field3'),
#                 }))
            
#             self.one2many_line_ids = lines

#         except Exception as e:
#             raise ValueError(f"Error reading the Excel file: {e}")



from odoo import models, fields,_
import openpyxl
import base64
from io import BytesIO
from odoo.exceptions import UserError

class AuditAccountImportWizard(models.TransientModel):
    _name = 'audit.account.import.wizard'
    _description = 'Import Wizard'

    excel_file = fields.Binary(string="Upload Excel File", required=True)
    excel_filename = fields.Char(string="File Name")

    def action_import_excel(self):
        if not self.excel_file:
            raise ValueError("Please upload a valid Excel file.")
        try:
            # Decode the uploaded file
            data = base64.b64decode(self.excel_file)
            file_stream = BytesIO(data)
            workbook = openpyxl.load_workbook(file_stream)
            sheet = workbook.active 
            # sheet = workbook.data
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                self.env['audit.account.account.line'].create({
                    'name': row[0].value,
                    'code': row[1].value,
                    'account_type': row[2].value,
                    'opening_debit': row[3].value,
                    'opening_credit': row[4].value,
                    'opening_balance': row[5].value,
                })
        except Exception as e:
            raise ValueError(f"Error processing Excel file: {e}")

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }




    # def action_import_excel(self):
    #     if not self.excel_file:
    #         raise ValueError("Please upload a valid Excel file.")

    #     try:
    #         # Decode the uploaded file
    #         data = base64.b64decode(self.excel_file)
    #         # Use BytesIO to convert the bytes to a file-like object
    #         file_stream = BytesIO(data)
    #         workbook = openpyxl.load_workbook(file_stream)  # Pass the BytesIO stream
    #         sheet = workbook.active  # Get the first sheet

    #         # Iterate over the rows (skip header)
    #         for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
    #             self.env['audit.account.account.line'].create({
    #                 'name': row[0].value,  # Adjust based on your column structure
    #                 'code': row[1].value,
    #                 'account_type': row[2].value,
    #                 'opening_debit': row[3].value,
    #                 'opening_credit': row[4].value,
    #                 'opening_balance': row[5].value,
    #             })
    #     except Exception as e:
    #         raise ValueError(f"Error processing Excel file: {e}")

    #     return {
    #         'type': 'ir.actions.client',
    #         'tag': 'reload',
    #     }